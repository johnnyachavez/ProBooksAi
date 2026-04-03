"""
ProBooks+ai – Accounting Software Workbook Generator
======================================================
Generates a fully-featured Excel workbook (.xlsx) that covers:

  1.  Dashboard          – Key metrics at a glance
  2.  Settings           – Company info, accounting method (Cash / Accrual)
  3.  Chart of Accounts  – Full COA with account types and normal balances
  4.  Customers          – Customer master database template
  5.  Vendors            – Vendor master database template
  6.  Invoices (AR)      – Invoice register with status tracking
  7.  Invoice Lines      – Line-item detail for each invoice
  8.  AR Payments        – Cash receipts against invoices
  9.  Bills (AP)         – Bill register with status tracking
  10. AP Payments        – Payments made against bills
  11. Journal Entries    – General ledger journal entry log
  12. Trial Balance      – Auto-aggregated debit / credit totals by account
  13. P&L Statement      – Profit & Loss (accrual or cash view)
  14. Balance Sheet      – Assets, Liabilities, Equity snapshot
  15. Cash Flow          – Operating / Investing / Financing sections
  16. AR Aging           – 0-30 / 31-60 / 61-90 / 90+ day buckets
  17. AP Aging           – Same aging buckets for payables

Run:
    python generate_workbook.py

The workbook is written to ``ProBooksAi_Accounting.xlsx`` (legacy default filename) in the current directory.

Install / CLI notes: ``README.md`` section **Excel workbook template (openpyxl)** (fragment ``#excel-workbook-template-openpyxl``).
"""

import os
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Colour / style palette
# ---------------------------------------------------------------------------
NAVY   = "1F3864"
BLUE   = "2E75B6"
LIGHT  = "D6E4F0"
GREEN  = "1E8449"
RED    = "C0392B"
AMBER  = "D4AC0D"
GREY   = "F2F2F2"
WHITE  = "FFFFFF"
BLACK  = "000000"

MONEY_FMT  = '#,##0.00'
DATE_FMT   = 'MM/DD/YYYY'
PCT_FMT    = '0.00%'

OUTPUT_FILE = "ProBooksAi_Accounting.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _font(bold=False, size=11, color=BLACK, italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _header_row(ws, row, cols, bg=NAVY, fg=WHITE, size=11, bold=True, height=22):
    """Write a styled header row."""
    for col, text in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.font      = _font(bold=bold, size=size, color=fg)
        c.fill      = _fill(bg)
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[row].height = height


def _data_row(ws, row, values, alt=False):
    """Write a plain data row with alternating colour."""
    bg = GREY if alt else WHITE
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = _fill(bg)
        c.alignment = _left()
        c.border    = _border()


def _money_cols(ws, col_indices, start_row, end_row):
    """Apply money format to a range of columns."""
    for ci in col_indices:
        col = get_column_letter(ci)
        for r in range(start_row, end_row + 1):
            ws[f"{col}{r}"].number_format = MONEY_FMT


def _date_cols(ws, col_indices, start_row, end_row):
    for ci in col_indices:
        col = get_column_letter(ci)
        for r in range(start_row, end_row + 1):
            ws[f"{col}{r}"].number_format = DATE_FMT


def _col_width(ws, widths):
    """Set column widths from a dict  {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _add_table(ws, name, ref, style="TableStyleMedium2"):
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True,
    )
    ws.add_table(tbl)


def _section_title(ws, row, text, col_span=2, bg=BLUE, fg=WHITE):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=col_span)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = _font(bold=True, size=12, color=fg)
    c.fill      = _fill(bg)
    c.alignment = _center()
    c.border    = _border()
    ws.row_dimensions[row].height = 20


def _kv(ws, row, label, value, label_col=1, val_col=2):
    lc = ws.cell(row=row, column=label_col, value=label)
    lc.font      = _font(bold=True, size=11)
    lc.fill      = _fill(LIGHT)
    lc.alignment = _left()
    lc.border    = _border()

    vc = ws.cell(row=row, column=val_col, value=value)
    vc.font      = _font(size=11)
    vc.fill      = _fill(WHITE)
    vc.alignment = _left()
    vc.border    = _border()
    return vc


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def _title_block(ws, title, subtitle=""):
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = title
    t.font      = _font(bold=True, size=18, color=WHITE)
    t.fill      = _fill(NAVY)
    t.alignment = _center()
    ws.row_dimensions[1].height = 36

    if subtitle:
        ws.merge_cells("A2:H2")
        s = ws["A2"]
        s.value     = subtitle
        s.font      = _font(italic=True, size=11, color=NAVY)
        s.fill      = _fill(LIGHT)
        s.alignment = _center()
        ws.row_dimensions[2].height = 18


# ---------------------------------------------------------------------------
# 1. Dashboard
# ---------------------------------------------------------------------------

def build_dashboard(ws, settings_name):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value     = "🏢  ProBooks+ai – Accounting Dashboard"
    t.font      = _font(bold=True, size=22, color=WHITE)
    t.fill      = _fill(NAVY)
    t.alignment = _center()
    ws.row_dimensions[1].height = 50

    # subtitle
    ws.merge_cells("A2:J2")
    s = ws["A2"]
    s.value     = f"Generated: {date.today().strftime('%B %d, %Y')}  |  All amounts in USD"
    s.font      = _font(italic=True, size=10, color=NAVY)
    s.fill      = _fill(LIGHT)
    s.alignment = _center()
    ws.row_dimensions[2].height = 18

    # ── KPI cards (row 4-5) ─────────────────────────────────────────────────
    kpi_labels = [
        "Total AR Outstanding", "Total AP Outstanding",
        "Cash Balance (est.)", "Net Income (YTD)",
    ]
    kpi_refs = [
        "=SUMIF(Invoices[Status],\"Open\",Invoices[Balance Due])",
        "=SUMIF(Bills[Status],\"Open\",Bills[Balance Due])",
        "=SUMPRODUCT((AR_Payments[Amount]))-SUMPRODUCT((AP_Payments[Amount]))",
        "=SUMIF('P&L'!B:B,\"Revenue\",'P&L'!C:C)-SUMIF('P&L'!B:B,\"Expense\",'P&L'!C:C)",
    ]
    col_map = [1, 3, 5, 7]
    for i, (label, ref) in enumerate(zip(kpi_labels, kpi_refs)):
        c = col_map[i]
        ws.merge_cells(start_row=4, start_column=c, end_row=4, end_column=c + 1)
        ws.merge_cells(start_row=5, start_column=c, end_row=5, end_column=c + 1)

        lbl = ws.cell(row=4, column=c, value=label)
        lbl.font      = _font(bold=True, size=10, color=WHITE)
        lbl.fill      = _fill(BLUE)
        lbl.alignment = _center()
        lbl.border    = _border()

        val = ws.cell(row=5, column=c, value=ref)
        val.font          = _font(bold=True, size=16, color=NAVY)
        val.fill          = _fill(LIGHT)
        val.alignment     = _center()
        val.border        = _border()
        val.number_format = MONEY_FMT

        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 36

    # ── Navigation table ────────────────────────────────────────────────────
    ws.merge_cells("A7:J7")
    nav = ws["A7"]
    nav.value     = "Sheet Navigator"
    nav.font      = _font(bold=True, size=13, color=WHITE)
    nav.fill      = _fill(BLUE)
    nav.alignment = _center()
    ws.row_dimensions[7].height = 22

    nav_items = [
        ("Settings",        "Company info & accounting method",      settings_name),
        ("Chart of Accounts","Full account list",                    "COA"),
        ("Customers",       "Customer master database",              "Customers"),
        ("Vendors",         "Vendor master database",                "Vendors"),
        ("Invoices",        "Accounts Receivable invoice register",  "Invoices"),
        ("Invoice Lines",   "Line-item detail per invoice",          "Invoice_Lines"),
        ("AR Payments",     "Cash receipts against invoices",        "AR_Payments"),
        ("Bills",           "Accounts Payable bill register",        "Bills"),
        ("AP Payments",     "Payments made against bills",           "AP_Payments"),
        ("Journal Entries", "General ledger journal entries",        "Journal_Entries"),
        ("Trial Balance",   "Aggregated debit/credit by account",    "Trial_Balance"),
        ("P&L",             "Profit & Loss Statement",               "P&L"),
        ("Balance Sheet",   "Assets / Liabilities / Equity",        "Balance_Sheet"),
        ("Cash Flow",       "Operating / Investing / Financing",     "Cash_Flow"),
        ("AR Aging",        "Accounts Receivable aging report",      "AR_Aging"),
        ("AP Aging",        "Accounts Payable aging report",         "AP_Aging"),
    ]
    _header_row(ws, 8, ["Sheet", "Description", "Tab Name"], bg=NAVY)
    for i, (sheet, desc, tab) in enumerate(nav_items, 9):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        for ci, val in enumerate([sheet, desc, tab], 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)
        ws.row_dimensions[i].height = 16

    _col_width(ws, {"A": 22, "B": 45, "C": 20, "D": 6, "E": 6, "F": 6, "G": 6, "H": 6})
    _freeze(ws, "A3")


# ---------------------------------------------------------------------------
# 2. Settings
# ---------------------------------------------------------------------------

def build_settings(ws):
    _title_block(ws, "⚙️  ProBooks+ai – Settings",
                 "Configure company information and accounting preferences")
    ws.sheet_view.showGridLines = False

    # ── Company Information ─────────────────────────────────────────────────
    _section_title(ws, 4, "Company Information", col_span=4, bg=NAVY)
    fields = [
        ("Company Name",        "Your Company, LLC"),
        ("Address Line 1",      "123 Main Street"),
        ("Address Line 2",      "Suite 100"),
        ("City, State ZIP",     "Anytown, CA 90210"),
        ("Phone",               "(555) 000-0000"),
        ("Email",               "accounting@yourcompany.com"),
        ("Website",             "www.yourcompany.com"),
        ("EIN / Tax ID",        "XX-XXXXXXX"),
        ("Fiscal Year Start",   "01/01"),
        ("Fiscal Year End",     "12/31"),
        ("Default Currency",    "USD"),
    ]
    for i, (label, val) in enumerate(fields, 5):
        _kv(ws, i, label, val, label_col=1, val_col=2)

    # ── Accounting Method ───────────────────────────────────────────────────
    _section_title(ws, 17, "Accounting Method", col_span=4, bg=NAVY)
    _kv(ws, 18, "Accounting Basis", "Accrual")
    ws["B18"].comment = None  # placeholder for a dropdown
    dv = DataValidation(type="list", formula1='"Accrual,Cash"', allow_blank=False)
    dv.sqref = "B18"
    ws.add_data_validation(dv)

    # ── Tax Settings ────────────────────────────────────────────────────────
    _section_title(ws, 20, "Tax Settings", col_span=4, bg=NAVY)
    tax_fields = [
        ("Default Sales Tax Rate (%)", 0.0725),
        ("Tax Account",               "Sales Tax Payable"),
    ]
    for i, (label, val) in enumerate(tax_fields, 21):
        vc = _kv(ws, i, label, val)
        if isinstance(val, float):
            vc.number_format = PCT_FMT

    # ── Invoice / Bill Defaults ─────────────────────────────────────────────
    _section_title(ws, 24, "Invoice / Bill Defaults", col_span=4, bg=NAVY)
    inv_fields = [
        ("Default Payment Terms (days)", 30),
        ("Next Invoice Number",          1001),
        ("Next Bill Number",             2001),
        ("Late Fee (%)",                 0.015),
        ("Default AR Account",           "Accounts Receivable"),
        ("Default AP Account",           "Accounts Payable"),
        ("Default Revenue Account",      "Sales Revenue"),
        ("Default COGS Account",         "Cost of Goods Sold"),
    ]
    for i, (label, val) in enumerate(inv_fields, 25):
        vc = _kv(ws, i, label, val)
        if isinstance(val, float):
            vc.number_format = PCT_FMT

    _col_width(ws, {"A": 32, "B": 35, "C": 5, "D": 5})
    _freeze(ws, "A3")


# ---------------------------------------------------------------------------
# 3. Chart of Accounts
# ---------------------------------------------------------------------------

COA_DATA = [
    # (Account #, Name, Type, Sub-Type, Normal Balance, Description)
    ("1000", "Cash – Checking",              "Asset",     "Current Asset",        "Debit",  "Primary operating bank account"),
    ("1010", "Cash – Savings",               "Asset",     "Current Asset",        "Debit",  "Business savings account"),
    ("1020", "Petty Cash",                   "Asset",     "Current Asset",        "Debit",  "On-hand petty cash fund"),
    ("1100", "Accounts Receivable",          "Asset",     "Current Asset",        "Debit",  "Amounts owed by customers"),
    ("1110", "Allowance for Doubtful Accts", "Asset",     "Current Asset",        "Credit", "Contra-AR – estimated uncollectible"),
    ("1200", "Inventory",                    "Asset",     "Current Asset",        "Debit",  "Goods held for sale"),
    ("1210", "Prepaid Expenses",             "Asset",     "Current Asset",        "Debit",  "Expenses paid in advance"),
    ("1500", "Equipment",                    "Asset",     "Fixed Asset",          "Debit",  "Machinery and equipment"),
    ("1510", "Accumulated Depreciation",     "Asset",     "Fixed Asset",          "Credit", "Contra-asset – accumulated depreciation"),
    ("1600", "Furniture & Fixtures",         "Asset",     "Fixed Asset",          "Debit",  "Office furniture"),
    ("1700", "Vehicles",                     "Asset",     "Fixed Asset",          "Debit",  "Company vehicles"),
    ("1800", "Leasehold Improvements",       "Asset",     "Fixed Asset",          "Debit",  "Improvements to leased space"),
    ("1900", "Other Assets",                 "Asset",     "Other Asset",          "Debit",  "Miscellaneous assets"),
    ("2000", "Accounts Payable",             "Liability", "Current Liability",    "Credit", "Amounts owed to vendors"),
    ("2100", "Accrued Liabilities",          "Liability", "Current Liability",    "Credit", "Expenses incurred but not yet paid"),
    ("2110", "Sales Tax Payable",            "Liability", "Current Liability",    "Credit", "Sales tax collected, not yet remitted"),
    ("2120", "Payroll Tax Payable",          "Liability", "Current Liability",    "Credit", "Payroll taxes withheld"),
    ("2200", "Short-Term Loans Payable",     "Liability", "Current Liability",    "Credit", "Loans due within 12 months"),
    ("2300", "Deferred Revenue",             "Liability", "Current Liability",    "Credit", "Cash received for services not yet delivered"),
    ("2500", "Long-Term Debt",               "Liability", "Long-Term Liability",  "Credit", "Loans/debt due after 12 months"),
    ("3000", "Common Stock",                 "Equity",    "Paid-In Capital",      "Credit", "Stockholder-contributed capital"),
    ("3100", "Retained Earnings",            "Equity",    "Retained Earnings",    "Credit", "Accumulated profits (prior periods)"),
    ("3200", "Owner's Draw",                 "Equity",    "Owner's Draw",         "Debit",  "Owner withdrawals"),
    ("4000", "Sales Revenue",                "Revenue",   "Operating Revenue",    "Credit", "Primary product/service revenue"),
    ("4100", "Service Revenue",              "Revenue",   "Operating Revenue",    "Credit", "Professional services income"),
    ("4200", "Other Income",                 "Revenue",   "Other Income",         "Credit", "Non-operating income"),
    ("4300", "Interest Income",              "Revenue",   "Other Income",         "Credit", "Interest earned"),
    ("5000", "Cost of Goods Sold",           "Expense",   "COGS",                 "Debit",  "Direct cost of products sold"),
    ("5100", "Direct Labor",                 "Expense",   "COGS",                 "Debit",  "Labor directly tied to production"),
    ("5200", "Freight & Shipping",           "Expense",   "COGS",                 "Debit",  "Shipping costs of goods sold"),
    ("6000", "Salaries & Wages",             "Expense",   "Operating Expense",    "Debit",  "Employee compensation"),
    ("6010", "Payroll Taxes",                "Expense",   "Operating Expense",    "Debit",  "Employer payroll taxes"),
    ("6020", "Employee Benefits",            "Expense",   "Operating Expense",    "Debit",  "Health insurance, 401k, etc."),
    ("6100", "Rent Expense",                 "Expense",   "Operating Expense",    "Debit",  "Office / facility rent"),
    ("6110", "Utilities",                    "Expense",   "Operating Expense",    "Debit",  "Electric, gas, water, internet"),
    ("6120", "Insurance",                    "Expense",   "Operating Expense",    "Debit",  "Business insurance premiums"),
    ("6200", "Advertising & Marketing",      "Expense",   "Operating Expense",    "Debit",  "Promotional expenses"),
    ("6210", "Software & Subscriptions",     "Expense",   "Operating Expense",    "Debit",  "SaaS tools and subscriptions"),
    ("6220", "Office Supplies",              "Expense",   "Operating Expense",    "Debit",  "Consumable office materials"),
    ("6300", "Travel & Entertainment",       "Expense",   "Operating Expense",    "Debit",  "Business travel and meals"),
    ("6310", "Vehicle Expense",              "Expense",   "Operating Expense",    "Debit",  "Gas, maintenance, mileage"),
    ("6400", "Professional Fees",            "Expense",   "Operating Expense",    "Debit",  "Legal, accounting, consulting"),
    ("6500", "Depreciation Expense",         "Expense",   "Operating Expense",    "Debit",  "Periodic depreciation charge"),
    ("6600", "Bad Debt Expense",             "Expense",   "Operating Expense",    "Debit",  "Write-offs of uncollectible receivables"),
    ("6700", "Bank Charges",                 "Expense",   "Operating Expense",    "Debit",  "Bank fees and service charges"),
    ("6800", "Interest Expense",             "Expense",   "Operating Expense",    "Debit",  "Interest paid on debt"),
    ("6900", "Miscellaneous Expense",        "Expense",   "Operating Expense",    "Debit",  "Other business expenses"),
    ("7000", "Income Tax Expense",           "Expense",   "Tax Expense",          "Debit",  "Corporate / business income taxes"),
]


def build_coa(ws):
    _title_block(ws, "📊  Chart of Accounts",
                 "All accounts used in the general ledger")
    ws.sheet_view.showGridLines = False

    cols = ["Account #", "Account Name", "Type", "Sub-Type",
            "Normal Balance", "Description", "Active"]
    _header_row(ws, 3, cols, bg=NAVY)

    for i, row in enumerate(COA_DATA, 4):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        data = list(row) + ["Yes"]
        for ci, val in enumerate(data, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 3 + len(COA_DATA)
    _add_table(ws, "COA", f"A3:{get_column_letter(7)}{last}")
    _col_width(ws, {"A": 12, "B": 30, "C": 12, "D": 22,
                    "E": 16, "F": 38, "G": 8})
    _freeze(ws, "A4")


# ---------------------------------------------------------------------------
# 4. Customers
# ---------------------------------------------------------------------------

CUSTOMER_HEADERS = [
    "Customer ID", "Company Name", "Contact Name", "Title",
    "Address Line 1", "Address Line 2", "City", "State", "ZIP", "Country",
    "Phone", "Mobile", "Email", "Website",
    "Payment Terms (days)", "Credit Limit", "AR Account",
    "Currency", "Tax Exempt?", "Tax ID", "Notes", "Active",
]

SAMPLE_CUSTOMERS = [
    ("C-1001","Acme Corporation","John Smith","CFO",
     "100 Industrial Blvd","","Springfield","IL","62701","USA",
     "217-555-0100","217-555-0101","jsmith@acme.com","www.acme.com",
     30,50000,"Accounts Receivable","USD","No","","Key account","Yes"),
    ("C-1002","Globex Ltd","Jane Doe","Controller",
     "200 Commerce St","Suite 5","Shelbyville","IL","62705","USA",
     "217-555-0200","","jdoe@globex.com","www.globex.com",
     45,25000,"Accounts Receivable","USD","No","","","Yes"),
    ("C-1003","Mom & Pop Shop","Bob Pop","Owner",
     "300 Main St","","Capital City","IL","62703","USA",
     "217-555-0300","217-555-0301","bob@momandpop.com","",
     15,5000,"Accounts Receivable","USD","No","","Cash preferred","Yes"),
]


def build_customers(ws):
    _title_block(ws, "👥  Customer Master Database",
                 "All customer records – source of truth for AR")
    ws.sheet_view.showGridLines = False

    _header_row(ws, 3, CUSTOMER_HEADERS, bg=NAVY)
    for i, row in enumerate(SAMPLE_CUSTOMERS, 4):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 3 + len(SAMPLE_CUSTOMERS)
    _add_table(ws, "Customers",
               f"A3:{get_column_letter(len(CUSTOMER_HEADERS))}{last}")

    widths = {"A":12,"B":25,"C":20,"D":15,"E":25,"F":15,"G":15,"H":8,
              "I":8,"J":10,"K":16,"L":16,"M":28,"N":22,"O":20,"P":14,
              "Q":20,"R":10,"S":12,"T":14,"U":25,"V":8}
    _col_width(ws, widths)
    _freeze(ws, "A4")


# ---------------------------------------------------------------------------
# 5. Vendors
# ---------------------------------------------------------------------------

VENDOR_HEADERS = [
    "Vendor ID", "Company Name", "Contact Name", "Title",
    "Address Line 1", "Address Line 2", "City", "State", "ZIP", "Country",
    "Phone", "Mobile", "Email", "Website",
    "Payment Terms (days)", "Account #",  "AP Account",
    "1099 Vendor?", "Tax ID", "Default Expense Account", "Notes", "Active",
]

SAMPLE_VENDORS = [
    ("V-2001","Office Depot","Sarah Lee","Sales Rep",
     "10 Supply Lane","","Chicago","IL","60601","USA",
     "312-555-0200","","slee@officedepot.com","www.officedepot.com",
     30,"CUST-9876","Accounts Payable","No","","Office Supplies","","Yes"),
    ("V-2002","ABC Telecom","Mike Ring","Account Mgr",
     "50 Fiber Rd","","Chicago","IL","60602","USA",
     "312-555-0300","312-555-0301","mring@abctel.com","www.abctel.com",
     30,"ACC-1234","Accounts Payable","No","","Utilities","Monthly invoice","Yes"),
    ("V-2003","Smith Consulting LLC","Alex Smith","Principal",
     "75 Consultant Dr","Suite 10","Springfield","IL","62701","USA",
     "217-555-0400","","asmith@smithco.com","www.smithco.com",
     15,"","Accounts Payable","Yes","35-1234567","Professional Fees","1099 contractor","Yes"),
]


def build_vendors(ws):
    _title_block(ws, "🏭  Vendor Master Database",
                 "All vendor records – source of truth for AP")
    ws.sheet_view.showGridLines = False

    _header_row(ws, 3, VENDOR_HEADERS, bg=NAVY)
    for i, row in enumerate(SAMPLE_VENDORS, 4):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 3 + len(SAMPLE_VENDORS)
    _add_table(ws, "Vendors",
               f"A3:{get_column_letter(len(VENDOR_HEADERS))}{last}")

    widths = {"A":12,"B":25,"C":20,"D":15,"E":25,"F":15,"G":15,"H":8,
              "I":8,"J":10,"K":16,"L":16,"M":28,"N":22,"O":20,"P":16,
              "Q":20,"R":14,"S":14,"T":22,"U":25,"V":8}
    _col_width(ws, widths)
    _freeze(ws, "A4")


# ---------------------------------------------------------------------------
# 6. Invoices (AR)
# ---------------------------------------------------------------------------

INVOICE_HEADERS = [
    "Invoice #", "Invoice Date", "Due Date",
    "Customer ID", "Customer Name", "PO Number",
    "Subtotal", "Tax Rate", "Tax Amount", "Total",
    "Payments Applied", "Balance Due",
    "Status", "AR Account", "Revenue Account",
    "Accounting Basis", "Notes",
]

TODAY        = date.today()
SAMPLE_INVOICES = [
    (1001, TODAY - timedelta(days=45), TODAY - timedelta(days=15),
     "C-1001","Acme Corporation","PO-500",
     5000.00, 0.0725, 362.50, 5362.50, 5362.50, 0.00,
     "Paid","Accounts Receivable","Sales Revenue","Accrual",""),
    (1002, TODAY - timedelta(days=30), TODAY,
     "C-1002","Globex Ltd","",
     2500.00, 0.0725, 181.25, 2681.25, 0.00, 2681.25,
     "Open","Accounts Receivable","Service Revenue","Accrual",""),
    (1003, TODAY - timedelta(days=10), TODAY + timedelta(days=20),
     "C-1001","Acme Corporation","PO-501",
     7500.00, 0.0725, 543.75, 8043.75, 0.00, 8043.75,
     "Open","Accounts Receivable","Sales Revenue","Accrual","Rush order"),
    (1004, TODAY - timedelta(days=60), TODAY - timedelta(days=30),
     "C-1003","Mom & Pop Shop","",
     800.00, 0.00, 0.00, 800.00, 0.00, 800.00,
     "Overdue","Accounts Receivable","Sales Revenue","Cash","Tax exempt"),
]

def build_invoices(ws):
    _title_block(ws, "🧾  Invoices – Accounts Receivable Register",
                 "Track all customer invoices and their payment status")
    ws.sheet_view.showGridLines = False

    _header_row(ws, 3, INVOICE_HEADERS, bg=NAVY)
    for i, row in enumerate(SAMPLE_INVOICES, 4):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 3 + len(SAMPLE_INVOICES)
    _add_table(ws, "Invoices",
               f"A3:{get_column_letter(len(INVOICE_HEADERS))}{last}")

    # Format money cols (G, I, J, K, L) and date cols (B, C)
    money_ci = [7, 9, 10, 11, 12]
    date_ci  = [2, 3]
    _money_cols(ws, money_ci, 4, last)
    _date_cols(ws, date_ci, 4, last)
    # Tax rate
    for r in range(4, last + 1):
        ws[f"H{r}"].number_format = PCT_FMT

    # Status conditional formatting by color (manual for samples)
    status_colors = {"Paid": "C8E6C9", "Open": "FFF9C4", "Overdue": "FFCDD2", "Void": "E0E0E0"}
    for i, row in enumerate(SAMPLE_INVOICES, 4):
        status = row[12]
        col = get_column_letter(13)  # M = Status
        ws[f"{col}{i}"].fill = _fill(status_colors.get(status, WHITE))

    widths = {"A":12,"B":14,"C":14,"D":12,"E":22,"F":14,"G":14,"H":10,
              "I":12,"J":14,"K":16,"L":14,"M":10,"N":20,"O":20,"P":16,"Q":20}
    _col_width(ws, widths)
    _freeze(ws, "A4")


# ---------------------------------------------------------------------------
# 7. Invoice Lines
# ---------------------------------------------------------------------------

INV_LINE_HEADERS = [
    "Invoice #", "Line #", "Item Code", "Description",
    "Quantity", "Unit Price", "Line Total", "Revenue Account", "COGS Account",
]

SAMPLE_INV_LINES = [
    (1001, 1, "PROD-A", "Widget Pro (100 units)", 100, 40.00, 4000.00, "Sales Revenue", "Cost of Goods Sold"),
    (1001, 2, "SVC-01", "Installation Services",   5, 200.00, 1000.00, "Service Revenue", ""),
    (1002, 1, "SVC-02", "Monthly Support (5 hrs)", 5, 500.00, 2500.00, "Service Revenue", ""),
    (1003, 1, "PROD-A", "Widget Pro (150 units)",  150, 40.00, 6000.00, "Sales Revenue", "Cost of Goods Sold"),
    (1003, 2, "PROD-B", "Widget Lite (30 units)",   30, 50.00, 1500.00, "Sales Revenue", "Cost of Goods Sold"),
    (1004, 1, "PROD-C", "Basic Widget (40 units)",  40, 20.00, 800.00,  "Sales Revenue", "Cost of Goods Sold"),
]

def build_invoice_lines(ws):
    _title_block(ws, "📋  Invoice Line Items",
                 "Detailed line-item breakdown for each invoice")
    ws.sheet_view.showGridLines = False
    _header_row(ws, 3, INV_LINE_HEADERS, bg=NAVY)

    for i, row in enumerate(SAMPLE_INV_LINES, 4):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 3 + len(SAMPLE_INV_LINES)
    _add_table(ws, "Invoice_Lines",
               f"A3:{get_column_letter(len(INV_LINE_HEADERS))}{last}")
    _money_cols(ws, [6, 7], 4, last)
    widths = {"A":12,"B":8,"C":12,"D":35,"E":10,"F":12,"G":14,"H":20,"I":22}
    _col_width(ws, widths)
    _freeze(ws, "A4")


# ---------------------------------------------------------------------------
# 8. AR Payments
# ---------------------------------------------------------------------------

AR_PMT_HEADERS = [
    "Payment #", "Payment Date", "Invoice #", "Customer ID", "Customer Name",
    "Amount", "Payment Method", "Reference / Check #",
    "Deposit Account", "Notes",
]

SAMPLE_AR_PMTS = [
    ("PMT-AR-001", TODAY - timedelta(days=20), 1001, "C-1001","Acme Corporation",
     5362.50, "ACH", "ACH-20240310", "Cash – Checking", "Full payment"),
]

def build_ar_payments(ws):
    _title_block(ws, "💰  AR Payments – Cash Receipts Journal",
                 "All payments received from customers against invoices")
    ws.sheet_view.showGridLines = False
    _header_row(ws, 3, AR_PMT_HEADERS, bg=NAVY)

    for i, row in enumerate(SAMPLE_AR_PMTS, 4):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 3 + len(SAMPLE_AR_PMTS)
    _add_table(ws, "AR_Payments",
               f"A3:{get_column_letter(len(AR_PMT_HEADERS))}{last}")
    _money_cols(ws, [6], 4, last)
    _date_cols(ws, [2], 4, last)
    widths = {"A":16,"B":14,"C":12,"D":12,"E":22,"F":14,"G":18,"H":20,"I":18,"J":25}
    _col_width(ws, widths)
    _freeze(ws, "A4")


# ---------------------------------------------------------------------------
# 9. Bills (AP)
# ---------------------------------------------------------------------------

BILL_HEADERS = [
    "Bill #", "Vendor Bill #", "Bill Date", "Due Date",
    "Vendor ID", "Vendor Name",
    "Description", "Subtotal", "Tax Amount", "Total",
    "Payments Applied", "Balance Due",
    "Status", "AP Account", "Expense Account",
    "Accounting Basis", "Notes",
]

SAMPLE_BILLS = [
    (2001, "INV-7711", TODAY - timedelta(days=30), TODAY,
     "V-2001","Office Depot","Office supplies Q1",
     450.00, 0.00, 450.00, 450.00, 0.00,
     "Paid","Accounts Payable","Office Supplies","Accrual",""),
    (2002, "INV-8899", TODAY - timedelta(days=15), TODAY + timedelta(days=15),
     "V-2002","ABC Telecom","Internet & phone – March",
     250.00, 0.00, 250.00, 0.00, 250.00,
     "Open","Accounts Payable","Utilities","Accrual",""),
    (2003, "INV-5500", TODAY - timedelta(days=45), TODAY - timedelta(days=15),
     "V-2003","Smith Consulting LLC","Strategy consulting Feb",
     3000.00, 0.00, 3000.00, 0.00, 3000.00,
     "Overdue","Accounts Payable","Professional Fees","Accrual","1099"),
]

def build_bills(ws):
    _title_block(ws, "📄  Bills – Accounts Payable Register",
                 "Track all vendor bills and their payment status")
    ws.sheet_view.showGridLines = False
    _header_row(ws, 3, BILL_HEADERS, bg=NAVY)

    for i, row in enumerate(SAMPLE_BILLS, 4):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 3 + len(SAMPLE_BILLS)
    _add_table(ws, "Bills",
               f"A3:{get_column_letter(len(BILL_HEADERS))}{last}")
    _money_cols(ws, [8, 9, 10, 11, 12], 4, last)
    _date_cols(ws, [3, 4], 4, last)

    status_colors = {"Paid": "C8E6C9", "Open": "FFF9C4", "Overdue": "FFCDD2", "Void": "E0E0E0"}
    for i, row in enumerate(SAMPLE_BILLS, 4):
        status = row[12]
        col = get_column_letter(13)
        ws[f"{col}{i}"].fill = _fill(status_colors.get(status, WHITE))

    widths = {"A":12,"B":16,"C":14,"D":14,"E":12,"F":22,"G":30,"H":14,
              "I":12,"J":14,"K":16,"L":14,"M":10,"N":20,"O":22,"P":16,"Q":20}
    _col_width(ws, widths)
    _freeze(ws, "A4")


# ---------------------------------------------------------------------------
# 10. AP Payments
# ---------------------------------------------------------------------------

AP_PMT_HEADERS = [
    "Payment #", "Payment Date", "Bill #", "Vendor ID", "Vendor Name",
    "Amount", "Payment Method", "Check # / Reference",
    "Payment Account", "Notes",
]

SAMPLE_AP_PMTS = [
    ("PMT-AP-001", TODAY - timedelta(days=25), 2001, "V-2001","Office Depot",
     450.00, "Check", "CHK-1001", "Cash – Checking", ""),
]

def build_ap_payments(ws):
    _title_block(ws, "💳  AP Payments – Cash Disbursements Journal",
                 "All payments made to vendors against bills")
    ws.sheet_view.showGridLines = False
    _header_row(ws, 3, AP_PMT_HEADERS, bg=NAVY)

    for i, row in enumerate(SAMPLE_AP_PMTS, 4):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 3 + len(SAMPLE_AP_PMTS)
    _add_table(ws, "AP_Payments",
               f"A3:{get_column_letter(len(AP_PMT_HEADERS))}{last}")
    _money_cols(ws, [6], 4, last)
    _date_cols(ws, [2], 4, last)
    widths = {"A":16,"B":14,"C":12,"D":12,"E":22,"F":14,"G":18,"H":20,"I":18,"J":25}
    _col_width(ws, widths)
    _freeze(ws, "A4")


# ---------------------------------------------------------------------------
# 11. Journal Entries
# ---------------------------------------------------------------------------

JE_HEADERS = [
    "Entry #", "Date", "Period", "Reference", "Description",
    "Account #", "Account Name", "Debit", "Credit",
    "Memo", "Entered By",
]

# Build sample JEs that mirror the sample invoices / payments
SAMPLE_JES = [
    # Invoice 1001 – record revenue (accrual)
    ("JE-001", TODAY-timedelta(45), "Mar", "INV-1001", "Invoice 1001 – Acme Corp revenue",
     "1100","Accounts Receivable", 5362.50, 0, "AR accrual", "System"),
    ("JE-001", TODAY-timedelta(45), "Mar", "INV-1001", "Invoice 1001 – Acme Corp revenue",
     "4000","Sales Revenue", 0, 4000.00, "Widget Pro 100 units", "System"),
    ("JE-001", TODAY-timedelta(45), "Mar", "INV-1001", "Invoice 1001 – Acme Corp revenue",
     "4100","Service Revenue", 0, 1000.00, "Installation Services", "System"),
    ("JE-001", TODAY-timedelta(45), "Mar", "INV-1001", "Invoice 1001 – Acme Corp revenue",
     "2110","Sales Tax Payable", 0, 362.50, "Sales tax", "System"),
    # Payment received on 1001
    ("JE-002", TODAY-timedelta(20), "Mar", "PMT-AR-001", "AR Payment – Acme Corp",
     "1000","Cash – Checking", 5362.50, 0, "ACH receipt", "System"),
    ("JE-002", TODAY-timedelta(20), "Mar", "PMT-AR-001", "AR Payment – Acme Corp",
     "1100","Accounts Receivable", 0, 5362.50, "", "System"),
    # Invoice 1002 – record revenue
    ("JE-003", TODAY-timedelta(30), "Mar", "INV-1002", "Invoice 1002 – Globex Ltd",
     "1100","Accounts Receivable", 2681.25, 0, "", "System"),
    ("JE-003", TODAY-timedelta(30), "Mar", "INV-1002", "Invoice 1002 – Globex Ltd",
     "4100","Service Revenue", 0, 2500.00, "", "System"),
    ("JE-003", TODAY-timedelta(30), "Mar", "INV-1002", "Invoice 1002 – Globex Ltd",
     "2110","Sales Tax Payable", 0, 181.25, "", "System"),
    # Bill 2001 – record expense
    ("JE-004", TODAY-timedelta(30), "Mar", "BILL-2001", "Bill 2001 – Office Depot supplies",
     "6220","Office Supplies", 450.00, 0, "", "System"),
    ("JE-004", TODAY-timedelta(30), "Mar", "BILL-2001", "Bill 2001 – Office Depot supplies",
     "2000","Accounts Payable", 0, 450.00, "", "System"),
    # AP Payment 2001
    ("JE-005", TODAY-timedelta(25), "Mar", "PMT-AP-001", "AP Payment – Office Depot",
     "2000","Accounts Payable", 450.00, 0, "Chk 1001", "System"),
    ("JE-005", TODAY-timedelta(25), "Mar", "PMT-AP-001", "AP Payment – Office Depot",
     "1000","Cash – Checking", 0, 450.00, "", "System"),
]

def build_journal_entries(ws):
    _title_block(ws, "📓  General Ledger – Journal Entries",
                 "Double-entry bookkeeping journal; every debit must equal its credit")
    ws.sheet_view.showGridLines = False
    _header_row(ws, 3, JE_HEADERS, bg=NAVY)

    for i, row in enumerate(SAMPLE_JES, 4):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 3 + len(SAMPLE_JES)
    _add_table(ws, "Journal_Entries",
               f"A3:{get_column_letter(len(JE_HEADERS))}{last}")
    _money_cols(ws, [8, 9], 4, last)
    _date_cols(ws, [2], 4, last)

    # Totals row
    tot_row = last + 2
    ws.cell(row=tot_row, column=1, value="TOTALS").font = _font(bold=True)
    ws.cell(row=tot_row, column=8,
            value=f"=SUM(H4:H{last})").number_format = MONEY_FMT
    ws.cell(row=tot_row, column=9,
            value=f"=SUM(I4:I{last})").number_format = MONEY_FMT
    ws.cell(row=tot_row, column=8).font = _font(bold=True)
    ws.cell(row=tot_row, column=9).font = _font(bold=True)

    # Balance check
    ws.cell(row=tot_row + 1, column=1,
            value="BALANCE CHECK (should be 0)").font = _font(bold=True, color=GREEN)
    chk = ws.cell(row=tot_row + 1, column=8,
                  value=f"=H{tot_row}-I{tot_row}")
    chk.number_format = MONEY_FMT
    chk.font = _font(bold=True, color=GREEN)

    widths = {"A":12,"B":14,"C":8,"D":16,"E":38,"F":12,"G":25,
              "H":14,"I":14,"J":25,"K":15}
    _col_width(ws, widths)
    _freeze(ws, "A4")


# ---------------------------------------------------------------------------
# 12. Trial Balance
# ---------------------------------------------------------------------------

def build_trial_balance(ws):
    _title_block(ws, "⚖️  Trial Balance",
                 "Aggregated debit and credit totals by account from Journal Entries")
    ws.sheet_view.showGridLines = False

    # Instructions
    ws.merge_cells("A3:F3")
    instr = ws["A3"]
    instr.value = (
        "This sheet pulls totals from Journal_Entries using SUMIF. "
        "Totals should balance (Total Debits = Total Credits)."
    )
    instr.font      = _font(italic=True, size=10, color=NAVY)
    instr.fill      = _fill(LIGHT)
    instr.alignment = _left()

    _header_row(ws, 4, ["Account #", "Account Name", "Type",
                         "Total Debits", "Total Credits", "Net Balance"], bg=NAVY)

    # Pull from COA list; add SUMIF formulas
    for i, (acct_num, acct_name, acct_type, *_) in enumerate(COA_DATA, 5):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        acct_col = get_column_letter(6)   # F = Account # in JE table

        row_vals = [acct_num, acct_name, acct_type,
                    f'=SUMIF(Journal_Entries[Account #],"{acct_num}",Journal_Entries[Debit])',
                    f'=SUMIF(Journal_Entries[Account #],"{acct_num}",Journal_Entries[Credit])',
                    f"=D{i}-E{i}"]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 4 + len(COA_DATA)
    _money_cols(ws, [4, 5, 6], 5, last)

    # Grand totals
    tot = last + 2
    ws.cell(row=tot, column=1, value="GRAND TOTAL").font = _font(bold=True)
    for ci, col in [(4,"D"),(5,"E"),(6,"F")]:
        c = ws.cell(row=tot, column=ci, value=f"=SUM({col}5:{col}{last})")
        c.number_format = MONEY_FMT
        c.font = _font(bold=True)

    # Balance check
    ws.cell(row=tot+1, column=1,
            value="CHECK (D-E should = 0)").font = _font(bold=True, color=GREEN)
    chk = ws.cell(row=tot+1, column=4,
                  value=f"=D{tot}-E{tot}")
    chk.number_format = MONEY_FMT
    chk.font = _font(bold=True, color=GREEN)

    widths = {"A":12,"B":30,"C":12,"D":16,"E":16,"F":16}
    _col_width(ws, widths)
    _freeze(ws, "A5")


# ---------------------------------------------------------------------------
# 13. P&L Statement
# ---------------------------------------------------------------------------

def build_pl(ws):
    _title_block(ws, "📈  Profit & Loss Statement",
                 f"For the period ending {TODAY.strftime('%B %d, %Y')}")
    ws.sheet_view.showGridLines = False

    # Hidden helper column B carries the section tag used by Dashboard formula
    # Col A = label, Col B = section type, Col C = amount

    def _pl_section(start_row, title, accounts, section_type, total_label):
        _section_title(ws, start_row, title, col_span=3, bg=NAVY)
        r = start_row + 1
        for acct_num, acct_name in accounts:
            lc = ws.cell(row=r, column=1, value=acct_name)
            lc.font = _font(size=10)
            lc.fill = _fill(GREY if r % 2 == 0 else WHITE)
            lc.alignment = _left()
            lc.border = _border()

            ws.cell(row=r, column=2, value=section_type).font = _font(size=9, color="AAAAAA")

            vc = ws.cell(
                row=r, column=3,
                value=f'=SUMIF(Journal_Entries[Account #],"{acct_num}",Journal_Entries[Credit])'
                      f'-SUMIF(Journal_Entries[Account #],"{acct_num}",Journal_Entries[Debit])',
            )
            vc.number_format = MONEY_FMT
            vc.font = _font(size=10)
            vc.fill = _fill(GREY if r % 2 == 0 else WHITE)
            vc.alignment = _left()
            vc.border = _border()
            r += 1

        # Subtotal
        sub_r = r
        sub = ws.cell(row=sub_r, column=1, value=total_label)
        sub.font = _font(bold=True)
        sub.fill = _fill(LIGHT)
        sub.border = _border()

        tot_c = ws.cell(row=sub_r, column=3,
                        value=f"=SUM(C{start_row+1}:C{sub_r-1})")
        tot_c.number_format = MONEY_FMT
        tot_c.font = _font(bold=True)
        tot_c.fill = _fill(LIGHT)
        tot_c.border = _border()
        return sub_r

    # Revenue
    rev_accounts = [(a[0], a[1]) for a in COA_DATA if a[2] == "Revenue"]
    rev_end = _pl_section(4, "REVENUE", rev_accounts, "Revenue", "Total Revenue")

    # COGS
    cogs_accounts = [(a[0], a[1]) for a in COA_DATA if a[3] == "COGS"]
    cogs_start = rev_end + 2
    cogs_end = _pl_section(cogs_start, "COST OF GOODS SOLD", cogs_accounts, "Expense", "Total COGS")

    # Gross Profit
    gp_row = cogs_end + 1
    gp = ws.cell(row=gp_row, column=1, value="GROSS PROFIT")
    gp.font = _font(bold=True, size=12, color=WHITE)
    gp.fill = _fill(GREEN)
    gp.border = _border()
    gpc = ws.cell(row=gp_row, column=3, value=f"=C{rev_end}-C{cogs_end}")
    gpc.number_format = MONEY_FMT
    gpc.font = _font(bold=True, size=12, color=WHITE)
    gpc.fill = _fill(GREEN)
    gpc.border = _border()

    # Operating Expenses
    op_accounts = [(a[0], a[1]) for a in COA_DATA
                   if a[2] == "Expense" and a[3] == "Operating Expense"]
    op_start = gp_row + 2
    op_end = _pl_section(op_start, "OPERATING EXPENSES", op_accounts, "Expense", "Total Operating Expenses")

    # Net Income
    ni_row = op_end + 1
    ni = ws.cell(row=ni_row, column=1, value="NET INCOME / (LOSS)")
    ni.font = _font(bold=True, size=13, color=WHITE)
    ni.fill = _fill(NAVY)
    ni.border = _border()
    nic = ws.cell(row=ni_row, column=3, value=f"=C{gp_row}-C{op_end}")
    nic.number_format = MONEY_FMT
    nic.font = _font(bold=True, size=13, color=WHITE)
    nic.fill = _fill(NAVY)
    nic.border = _border()

    # Col B hidden helper (width 0)
    _col_width(ws, {"A": 35, "B": 0, "C": 18})
    _freeze(ws, "A3")


# ---------------------------------------------------------------------------
# 14. Balance Sheet
# ---------------------------------------------------------------------------

def build_balance_sheet(ws):
    _title_block(ws, "📊  Balance Sheet",
                 f"As of {TODAY.strftime('%B %d, %Y')}")
    ws.sheet_view.showGridLines = False

    def _bs_section(start_row, title, accounts, normal_debit=True, bg_title=NAVY):
        _section_title(ws, start_row, title, col_span=3, bg=bg_title)
        r = start_row + 1
        for acct_num, acct_name in accounts:
            lc = ws.cell(row=r, column=1, value=acct_name)
            lc.font = _font(size=10)
            lc.fill = _fill(GREY if r % 2 == 0 else WHITE)
            lc.alignment = _left()
            lc.border = _border()

            # Normal debit accounts: net = debits - credits
            if normal_debit:
                formula = (
                    f'=SUMIF(Journal_Entries[Account #],"{acct_num}",Journal_Entries[Debit])'
                    f'-SUMIF(Journal_Entries[Account #],"{acct_num}",Journal_Entries[Credit])'
                )
            else:
                formula = (
                    f'=SUMIF(Journal_Entries[Account #],"{acct_num}",Journal_Entries[Credit])'
                    f'-SUMIF(Journal_Entries[Account #],"{acct_num}",Journal_Entries[Debit])'
                )
            vc = ws.cell(row=r, column=3, value=formula)
            vc.number_format = MONEY_FMT
            vc.font = _font(size=10)
            vc.fill = _fill(GREY if r % 2 == 0 else WHITE)
            vc.alignment = _left()
            vc.border = _border()
            r += 1

        sub_r = r
        sub = ws.cell(row=sub_r, column=1, value=f"Total {title}")
        sub.font = _font(bold=True)
        sub.fill = _fill(LIGHT)
        sub.border = _border()
        tot = ws.cell(row=sub_r, column=3,
                      value=f"=SUM(C{start_row+1}:C{sub_r-1})")
        tot.number_format = MONEY_FMT
        tot.font = _font(bold=True)
        tot.fill = _fill(LIGHT)
        tot.border = _border()
        return sub_r

    # Assets
    cur_assets  = [(a[0],a[1]) for a in COA_DATA if a[3]=="Current Asset"]
    fix_assets  = [(a[0],a[1]) for a in COA_DATA if a[3]=="Fixed Asset"]
    other_assets = [(a[0],a[1]) for a in COA_DATA if a[3]=="Other Asset"]

    _section_title(ws, 4, "ASSETS", col_span=3, bg=BLUE)

    cur_end  = _bs_section(5,  "Current Assets",   cur_assets,  True, BLUE)
    fix_end  = _bs_section(cur_end+2,  "Fixed Assets",     fix_assets,  True, BLUE)
    oth_end  = _bs_section(fix_end+2,  "Other Assets",     other_assets,True, BLUE)

    ta_row = oth_end + 1
    ta = ws.cell(row=ta_row, column=1, value="TOTAL ASSETS")
    ta.font = _font(bold=True, size=12, color=WHITE)
    ta.fill = _fill(NAVY)
    ta.border = _border()
    tac = ws.cell(row=ta_row, column=3,
                  value=f"=C{cur_end}+C{fix_end}+C{oth_end}")
    tac.number_format = MONEY_FMT
    tac.font = _font(bold=True, size=12, color=WHITE)
    tac.fill = _fill(NAVY)
    tac.border = _border()

    # Liabilities
    cur_liab  = [(a[0],a[1]) for a in COA_DATA if a[3]=="Current Liability"]
    lt_liab   = [(a[0],a[1]) for a in COA_DATA if a[3]=="Long-Term Liability"]

    liab_start = ta_row + 2
    _section_title(ws, liab_start, "LIABILITIES", col_span=3, bg=RED)
    cl_end = _bs_section(liab_start+1, "Current Liabilities", cur_liab, False, RED)
    ll_end = _bs_section(cl_end+2,     "Long-Term Liabilities", lt_liab, False, RED)

    tl_row = ll_end + 1
    tl = ws.cell(row=tl_row, column=1, value="TOTAL LIABILITIES")
    tl.font = _font(bold=True, size=12, color=WHITE)
    tl.fill = _fill(RED)
    tl.border = _border()
    tlc = ws.cell(row=tl_row, column=3,
                  value=f"=C{cl_end}+C{ll_end}")
    tlc.number_format = MONEY_FMT
    tlc.font = _font(bold=True, size=12, color=WHITE)
    tlc.fill = _fill(RED)
    tlc.border = _border()

    # Equity
    eq_accts = [(a[0],a[1]) for a in COA_DATA if a[2]=="Equity"]
    eq_start = tl_row + 2
    _section_title(ws, eq_start, "EQUITY", col_span=3, bg=GREEN)
    eq_end = _bs_section(eq_start+1, "Owner's Equity / Stockholders' Equity",
                         eq_accts, False, GREEN)

    # Net Income flows into equity
    ni_eq_row = eq_end + 1
    ni_eq = ws.cell(row=ni_eq_row, column=1, value="Net Income (Current Period)")
    ni_eq.font = _font(bold=True)
    ni_eq.fill = _fill(LIGHT)
    ni_eq.border = _border()
    # Link to P&L net income row
    ni_eq_c = ws.cell(row=ni_eq_row, column=3, value="='P&L'!C1")  # approximate
    ni_eq_c.number_format = MONEY_FMT
    ni_eq_c.font = _font(bold=True)
    ni_eq_c.fill = _fill(LIGHT)
    ni_eq_c.border = _border()

    te_row = ni_eq_row + 1
    te = ws.cell(row=te_row, column=1, value="TOTAL EQUITY")
    te.font = _font(bold=True, size=12, color=WHITE)
    te.fill = _fill(GREEN)
    te.border = _border()
    tec = ws.cell(row=te_row, column=3,
                  value=f"=C{eq_end}+C{ni_eq_row}")
    tec.number_format = MONEY_FMT
    tec.font = _font(bold=True, size=12, color=WHITE)
    tec.fill = _fill(GREEN)
    tec.border = _border()

    # Balance check
    ttle_row = te_row + 2
    ttl = ws.cell(row=ttle_row, column=1, value="TOTAL LIABILITIES + EQUITY")
    ttl.font = _font(bold=True, size=12)
    ttl.fill = _fill(LIGHT)
    ttl.border = _border()
    ttlc = ws.cell(row=ttle_row, column=3,
                   value=f"=C{tl_row}+C{te_row}")
    ttlc.number_format = MONEY_FMT
    ttlc.font = _font(bold=True, size=12)
    ttlc.fill = _fill(LIGHT)
    ttlc.border = _border()

    chk_row = ttle_row + 1
    chk = ws.cell(row=chk_row, column=1,
                  value="Balance Check (Assets - Liab - Equity) should = 0")
    chk.font = _font(bold=True, color=GREEN)
    chkc = ws.cell(row=chk_row, column=3,
                   value=f"=C{ta_row}-C{ttle_row}")
    chkc.number_format = MONEY_FMT
    chkc.font = _font(bold=True, color=GREEN)

    _col_width(ws, {"A": 40, "B": 0, "C": 18})
    _freeze(ws, "A3")


# ---------------------------------------------------------------------------
# 15. Cash Flow Statement
# ---------------------------------------------------------------------------

def build_cash_flow(ws):
    _title_block(ws, "💧  Cash Flow Statement",
                 f"For the period ending {TODAY.strftime('%B %d, %Y')}")
    ws.sheet_view.showGridLines = False

    row = 4

    def section(title, items, bg=BLUE):
        nonlocal row
        _section_title(ws, row, title, col_span=3, bg=bg)
        row += 1
        start = row
        for label, formula in items:
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = _font(size=10)
            lc.fill = _fill(GREY if row % 2 == 0 else WHITE)
            lc.alignment = _left()
            lc.border = _border()

            vc = ws.cell(row=row, column=3, value=formula)
            vc.number_format = MONEY_FMT
            vc.font = _font(size=10)
            vc.fill = _fill(GREY if row % 2 == 0 else WHITE)
            vc.alignment = _left()
            vc.border = _border()
            row += 1

        sub = ws.cell(row=row, column=1, value=f"Net Cash – {title}")
        sub.font = _font(bold=True)
        sub.fill = _fill(LIGHT)
        sub.border = _border()
        tot = ws.cell(row=row, column=3, value=f"=SUM(C{start}:C{row-1})")
        tot.number_format = MONEY_FMT
        tot.font = _font(bold=True)
        tot.fill = _fill(LIGHT)
        tot.border = _border()
        net_row = row
        row += 2
        return net_row

    operating_items = [
        ("Net Income",
         "=SUMIF(Journal_Entries[Account #],\"4000\",Journal_Entries[Credit])"
         "+SUMIF(Journal_Entries[Account #],\"4100\",Journal_Entries[Credit])"
         "-SUMIF(Journal_Entries[Account #],\"6000\",Journal_Entries[Debit])"
         "-SUMIF(Journal_Entries[Account #],\"6220\",Journal_Entries[Debit])"
         "-SUMIF(Journal_Entries[Account #],\"6100\",Journal_Entries[Debit])"),
        ("(Increase)/Decrease in Accounts Receivable",
         "=-(SUMIF(Journal_Entries[Account #],\"1100\",Journal_Entries[Debit])"
         "-SUMIF(Journal_Entries[Account #],\"1100\",Journal_Entries[Credit]))"),
        ("Increase/(Decrease) in Accounts Payable",
         "=SUMIF(Journal_Entries[Account #],\"2000\",Journal_Entries[Credit])"
         "-SUMIF(Journal_Entries[Account #],\"2000\",Journal_Entries[Debit])"),
        ("Increase/(Decrease) in Sales Tax Payable",
         "=SUMIF(Journal_Entries[Account #],\"2110\",Journal_Entries[Credit])"
         "-SUMIF(Journal_Entries[Account #],\"2110\",Journal_Entries[Debit])"),
        ("Add: Depreciation & Amortization",
         "=SUMIF(Journal_Entries[Account #],\"6500\",Journal_Entries[Debit])"),
    ]
    op_net = section("Operating Activities", operating_items, BLUE)

    investing_items = [
        ("Purchase of Equipment",
         "=-(SUMIF(Journal_Entries[Account #],\"1500\",Journal_Entries[Debit])"
         "-SUMIF(Journal_Entries[Account #],\"1500\",Journal_Entries[Credit]))"),
        ("Purchase of Furniture & Fixtures",
         "=-(SUMIF(Journal_Entries[Account #],\"1600\",Journal_Entries[Debit])"
         "-SUMIF(Journal_Entries[Account #],\"1600\",Journal_Entries[Credit]))"),
        ("Proceeds from Sale of Assets", "=0"),
    ]
    inv_net = section("Investing Activities", investing_items, AMBER)

    financing_items = [
        ("Proceeds from Loans",
         "=SUMIF(Journal_Entries[Account #],\"2500\",Journal_Entries[Credit])"
         "-SUMIF(Journal_Entries[Account #],\"2500\",Journal_Entries[Debit])"),
        ("Repayments of Loans",
         "=-(SUMIF(Journal_Entries[Account #],\"2200\",Journal_Entries[Debit])"
         "-SUMIF(Journal_Entries[Account #],\"2200\",Journal_Entries[Credit]))"),
        ("Owner Contributions / (Draws)",
         "=SUMIF(Journal_Entries[Account #],\"3000\",Journal_Entries[Credit])"
         "-SUMIF(Journal_Entries[Account #],\"3200\",Journal_Entries[Debit])"),
    ]
    fin_net = section("Financing Activities", financing_items, GREEN)

    # Net Change in Cash
    net_row = row
    nc = ws.cell(row=net_row, column=1, value="NET CHANGE IN CASH")
    nc.font = _font(bold=True, size=13, color=WHITE)
    nc.fill = _fill(NAVY)
    nc.border = _border()
    ncc = ws.cell(row=net_row, column=3,
                  value=f"=C{op_net}+C{inv_net}+C{fin_net}")
    ncc.number_format = MONEY_FMT
    ncc.font = _font(bold=True, size=13, color=WHITE)
    ncc.fill = _fill(NAVY)
    ncc.border = _border()

    _col_width(ws, {"A": 45, "B": 0, "C": 18})
    _freeze(ws, "A3")


# ---------------------------------------------------------------------------
# 16. AR Aging
# ---------------------------------------------------------------------------

def build_ar_aging(ws):
    _title_block(ws, "📅  AR Aging Report",
                 f"As of {TODAY.strftime('%B %d, %Y')} – Outstanding customer invoices by age")
    ws.sheet_view.showGridLines = False

    headers = ["Invoice #", "Customer", "Invoice Date", "Due Date",
               "Total", "Current", "1-30 Days",
               "31-60 Days", "61-90 Days", "91+ Days", "Status"]
    _header_row(ws, 3, headers, bg=NAVY)

    aging_data = []
    for inv in SAMPLE_INVOICES:
        inv_num, inv_date, due_date, cust_id, cust_name = inv[0], inv[1], inv[2], inv[3], inv[4]
        balance = inv[11]
        if balance == 0:
            continue
        days_overdue = (TODAY - due_date).days if isinstance(due_date, date) else 0

        cur = bal_1_30 = bal_31_60 = bal_61_90 = bal_91 = 0.0
        if days_overdue <= 0:
            cur = balance
        elif days_overdue <= 30:
            bal_1_30 = balance
        elif days_overdue <= 60:
            bal_31_60 = balance
        elif days_overdue <= 90:
            bal_61_90 = balance
        else:
            bal_91 = balance

        aging_data.append((inv_num, cust_name, inv_date, due_date,
                            balance, cur, bal_1_30, bal_31_60, bal_61_90, bal_91, inv[12]))

    for i, row in enumerate(aging_data, 4):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 3 + len(aging_data)
    if aging_data:
        _add_table(ws, "AR_Aging",
                   f"A3:{get_column_letter(len(headers))}{last}")
        _money_cols(ws, [5, 6, 7, 8, 9, 10], 4, last)
        _date_cols(ws, [3, 4], 4, last)

    # Totals
    tot_r = last + 2
    ws.cell(row=tot_r, column=1, value="TOTALS").font = _font(bold=True)
    for ci in [5, 6, 7, 8, 9, 10]:
        col = get_column_letter(ci)
        c = ws.cell(row=tot_r, column=ci, value=f"=SUM({col}4:{col}{last})")
        c.number_format = MONEY_FMT
        c.font = _font(bold=True)
        c.fill = _fill(LIGHT)
        c.border = _border()

    widths = {"A":12,"B":25,"C":14,"D":14,"E":14,"F":12,"G":12,"H":14,"I":14,"J":12,"K":10}
    _col_width(ws, widths)
    _freeze(ws, "A4")


# ---------------------------------------------------------------------------
# 17. AP Aging
# ---------------------------------------------------------------------------

def build_ap_aging(ws):
    _title_block(ws, "📅  AP Aging Report",
                 f"As of {TODAY.strftime('%B %d, %Y')} – Outstanding vendor bills by age")
    ws.sheet_view.showGridLines = False

    headers = ["Bill #", "Vendor", "Bill Date", "Due Date",
               "Total", "Current", "1-30 Days",
               "31-60 Days", "61-90 Days", "91+ Days", "Status"]
    _header_row(ws, 3, headers, bg=NAVY)

    aging_data = []
    for bill in SAMPLE_BILLS:
        bill_num, bill_date, due_date, vend_id, vend_name = (
            bill[0], bill[2], bill[3], bill[4], bill[5])
        balance = bill[11]
        if balance == 0:
            continue
        days_overdue = (TODAY - due_date).days if isinstance(due_date, date) else 0

        cur = bal_1_30 = bal_31_60 = bal_61_90 = bal_91 = 0.0
        if days_overdue <= 0:
            cur = balance
        elif days_overdue <= 30:
            bal_1_30 = balance
        elif days_overdue <= 60:
            bal_31_60 = balance
        elif days_overdue <= 90:
            bal_61_90 = balance
        else:
            bal_91 = balance

        aging_data.append((bill_num, vend_name, bill_date, due_date,
                            balance, cur, bal_1_30, bal_31_60, bal_61_90, bal_91, bill[12]))

    for i, row in enumerate(aging_data, 4):
        alt = (i % 2 == 0)
        bg  = GREY if alt else WHITE
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.alignment = _left()
            c.border    = _border()
            c.font      = _font(size=10)

    last = 3 + len(aging_data)
    if aging_data:
        _add_table(ws, "AP_Aging",
                   f"A3:{get_column_letter(len(headers))}{last}")
        _money_cols(ws, [5, 6, 7, 8, 9, 10], 4, last)
        _date_cols(ws, [3, 4], 4, last)

    # Totals
    tot_r = last + 2
    ws.cell(row=tot_r, column=1, value="TOTALS").font = _font(bold=True)
    for ci in [5, 6, 7, 8, 9, 10]:
        col = get_column_letter(ci)
        c = ws.cell(row=tot_r, column=ci, value=f"=SUM({col}4:{col}{last})")
        c.number_format = MONEY_FMT
        c.font = _font(bold=True)
        c.fill = _fill(LIGHT)
        c.border = _border()

    widths = {"A":12,"B":25,"C":14,"D":14,"E":14,"F":12,"G":12,"H":14,"I":14,"J":12,"K":10}
    _col_width(ws, widths)
    _freeze(ws, "A4")


# ---------------------------------------------------------------------------
# Main build function
# ---------------------------------------------------------------------------

def build_workbook(path=OUTPUT_FILE):
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    def add(title, tab_color=None):
        ws = wb.create_sheet(title=title)
        if tab_color:
            ws.sheet_properties.tabColor = tab_color
        return ws

    dash_ws  = add("Dashboard",      NAVY)
    sett_ws  = add("Settings",       "888888")
    coa_ws   = add("COA",            BLUE)
    cust_ws  = add("Customers",      "2196F3")
    vend_ws  = add("Vendors",        "FF9800")
    inv_ws   = add("Invoices",       GREEN)
    invl_ws  = add("Invoice_Lines",  "4CAF50")
    ar_pmt   = add("AR_Payments",    "8BC34A")
    bill_ws  = add("Bills",          RED)
    ap_pmt   = add("AP_Payments",    "F44336")
    je_ws    = add("Journal_Entries","9C27B0")
    tb_ws    = add("Trial_Balance",  AMBER)
    pl_ws    = add("P&L",            GREEN)
    bs_ws    = add("Balance_Sheet",  NAVY)
    cf_ws    = add("Cash_Flow",      "00BCD4")
    ar_age   = add("AR_Aging",       "FFC107")
    ap_age   = add("AP_Aging",       "FF5722")

    build_dashboard(dash_ws, "Settings")
    build_settings(sett_ws)
    build_coa(coa_ws)
    build_customers(cust_ws)
    build_vendors(vend_ws)
    build_invoices(inv_ws)
    build_invoice_lines(invl_ws)
    build_ar_payments(ar_pmt)
    build_bills(bill_ws)
    build_ap_payments(ap_pmt)
    build_journal_entries(je_ws)
    build_trial_balance(tb_ws)
    build_pl(pl_ws)
    build_balance_sheet(bs_ws)
    build_cash_flow(cf_ws)
    build_ar_aging(ar_age)
    build_ap_aging(ap_age)

    wb.save(path)
    print(f"✅  Workbook saved: {os.path.abspath(path)}")
    print(f"    Sheets: {[ws.title for ws in wb.worksheets]}")
    return path

