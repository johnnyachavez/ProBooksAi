"""
Tests for generate_workbook.py
Run with:  python -m pytest test_workbook.py -v
"""

import os
import tempfile

import pytest
from openpyxl import load_workbook

# Import the module under test
from generate_workbook import (
    COA_DATA,
    CUSTOMER_HEADERS,
    VENDOR_HEADERS,
    INVOICE_HEADERS,
    BILL_HEADERS,
    AR_PMT_HEADERS,
    AP_PMT_HEADERS,
    INV_LINE_HEADERS,
    JE_HEADERS,
    SAMPLE_CUSTOMERS,
    SAMPLE_VENDORS,
    SAMPLE_INVOICES,
    SAMPLE_BILLS,
    SAMPLE_AR_PMTS,
    SAMPLE_AP_PMTS,
    SAMPLE_INV_LINES,
    SAMPLE_JES,
    build_workbook,
)

EXPECTED_SHEETS = [
    "Dashboard",
    "Settings",
    "COA",
    "Customers",
    "Vendors",
    "Invoices",
    "Invoice_Lines",
    "AR_Payments",
    "Bills",
    "AP_Payments",
    "Journal_Entries",
    "Trial_Balance",
    "P&L",
    "Balance_Sheet",
    "Cash_Flow",
    "AR_Aging",
    "AP_Aging",
]


@pytest.fixture(scope="module")
def workbook_path(tmp_path_factory):
    """Generate workbook once per test session in a temp directory."""
    tmp = tmp_path_factory.mktemp("wb")
    path = str(tmp / "ProBooksAi_Accounting.xlsx")
    build_workbook(path)
    return path


@pytest.fixture(scope="module")
def wb(workbook_path):
    return load_workbook(workbook_path, data_only=False)


# ---------------------------------------------------------------------------
# Sheet presence
# ---------------------------------------------------------------------------

class TestSheets:
    def test_all_expected_sheets_present(self, wb):
        assert wb.sheetnames == EXPECTED_SHEETS

    def test_sheet_count(self, wb):
        assert len(wb.sheetnames) == len(EXPECTED_SHEETS)


# ---------------------------------------------------------------------------
# Chart of Accounts
# ---------------------------------------------------------------------------

class TestCOA:
    def test_coa_has_entries(self):
        assert len(COA_DATA) > 0

    def test_coa_covers_all_types(self):
        types = {row[2] for row in COA_DATA}
        assert types == {"Asset", "Liability", "Equity", "Revenue", "Expense"}

    def test_coa_normal_balance_values(self):
        balances = {row[4] for row in COA_DATA}
        assert balances == {"Debit", "Credit"}

    def test_coa_has_revenue_accounts(self):
        rev = [r for r in COA_DATA if r[2] == "Revenue"]
        assert len(rev) >= 3

    def test_coa_has_expense_accounts(self):
        exp = [r for r in COA_DATA if r[2] == "Expense"]
        assert len(exp) >= 10

    def test_coa_sheet_rows(self, wb):
        ws = wb["COA"]
        # Header row at 3, data from 4 onwards
        assert ws.cell(row=3, column=1).value == "Account #"
        # At least as many data rows as COA entries
        assert ws.max_row >= 3 + len(COA_DATA)

    def test_coa_first_account(self, wb):
        ws = wb["COA"]
        assert ws.cell(row=4, column=1).value == "1000"


# ---------------------------------------------------------------------------
# Customers
# ---------------------------------------------------------------------------

class TestCustomers:
    def test_customer_headers_count(self):
        assert len(CUSTOMER_HEADERS) == 22

    def test_sample_customers_exist(self):
        assert len(SAMPLE_CUSTOMERS) >= 3

    def test_customer_ids_unique(self):
        ids = [c[0] for c in SAMPLE_CUSTOMERS]
        assert len(ids) == len(set(ids))

    def test_customer_sheet_header(self, wb):
        ws = wb["Customers"]
        assert ws.cell(row=3, column=1).value == "Customer ID"
        assert ws.cell(row=3, column=2).value == "Company Name"

    def test_customer_sheet_has_data(self, wb):
        ws = wb["Customers"]
        assert ws.max_row >= 3 + len(SAMPLE_CUSTOMERS)


# ---------------------------------------------------------------------------
# Vendors
# ---------------------------------------------------------------------------

class TestVendors:
    def test_vendor_headers_count(self):
        assert len(VENDOR_HEADERS) == 22

    def test_sample_vendors_exist(self):
        assert len(SAMPLE_VENDORS) >= 3

    def test_vendor_ids_unique(self):
        ids = [v[0] for v in SAMPLE_VENDORS]
        assert len(ids) == len(set(ids))

    def test_vendor_sheet_header(self, wb):
        ws = wb["Vendors"]
        assert ws.cell(row=3, column=1).value == "Vendor ID"

    def test_vendor_sheet_has_data(self, wb):
        ws = wb["Vendors"]
        assert ws.max_row >= 3 + len(SAMPLE_VENDORS)


# ---------------------------------------------------------------------------
# Invoices (AR)
# ---------------------------------------------------------------------------

class TestInvoices:
    def test_invoice_headers_count(self):
        assert len(INVOICE_HEADERS) == 17

    def test_sample_invoices_exist(self):
        assert len(SAMPLE_INVOICES) >= 3

    def test_invoice_numbers_unique(self):
        nums = [i[0] for i in SAMPLE_INVOICES]
        assert len(nums) == len(set(nums))

    def test_invoice_balance_due_calculation(self):
        """Balance Due = Total - Payments Applied."""
        for inv in SAMPLE_INVOICES:
            total    = inv[9]
            payments = inv[10]
            balance  = inv[11]
            assert abs(balance - (total - payments)) < 0.01, (
                f"Invoice {inv[0]}: balance {balance} != total {total} - payments {payments}"
            )

    def test_invoice_tax_calculation(self):
        """Tax Amount = Subtotal * Tax Rate (within rounding)."""
        for inv in SAMPLE_INVOICES:
            sub      = inv[6]
            tax_rate = inv[7]
            tax_amt  = inv[8]
            expected = round(sub * tax_rate, 2)
            assert abs(tax_amt - expected) < 0.01, (
                f"Invoice {inv[0]}: tax {tax_amt} != {sub} * {tax_rate} = {expected}"
            )

    def test_invoice_total_calculation(self):
        """Total = Subtotal + Tax Amount."""
        for inv in SAMPLE_INVOICES:
            assert abs(inv[9] - (inv[6] + inv[8])) < 0.01

    def test_invoice_statuses(self):
        statuses = {i[12] for i in SAMPLE_INVOICES}
        assert statuses <= {"Open", "Paid", "Overdue", "Void"}

    def test_invoice_sheet_header(self, wb):
        ws = wb["Invoices"]
        assert ws.cell(row=3, column=1).value == "Invoice #"

    def test_invoice_sheet_data(self, wb):
        ws = wb["Invoices"]
        assert ws.max_row >= 3 + len(SAMPLE_INVOICES)


# ---------------------------------------------------------------------------
# Invoice Lines
# ---------------------------------------------------------------------------

class TestInvoiceLines:
    def test_inv_line_headers_count(self):
        assert len(INV_LINE_HEADERS) == 9

    def test_sample_lines_exist(self):
        assert len(SAMPLE_INV_LINES) >= 4

    def test_line_total_calculation(self):
        """Line Total = Quantity * Unit Price."""
        for line in SAMPLE_INV_LINES:
            qty   = line[4]
            price = line[5]
            total = line[6]
            assert abs(total - qty * price) < 0.01, (
                f"Line {line[2]}: {qty} * {price} != {total}"
            )

    def test_invoice_line_sheet(self, wb):
        ws = wb["Invoice_Lines"]
        assert ws.cell(row=3, column=1).value == "Invoice #"


# ---------------------------------------------------------------------------
# AR Payments
# ---------------------------------------------------------------------------

class TestARPayments:
    def test_ar_pmt_headers_count(self):
        assert len(AR_PMT_HEADERS) == 10

    def test_sample_ar_pmts_exist(self):
        assert len(SAMPLE_AR_PMTS) >= 1

    def test_ar_pmt_payment_ids_unique(self):
        ids = [p[0] for p in SAMPLE_AR_PMTS]
        assert len(ids) == len(set(ids))

    def test_ar_pmt_sheet(self, wb):
        ws = wb["AR_Payments"]
        assert ws.cell(row=3, column=1).value == "Payment #"


# ---------------------------------------------------------------------------
# Bills (AP)
# ---------------------------------------------------------------------------

class TestBills:
    def test_bill_headers_count(self):
        assert len(BILL_HEADERS) == 17

    def test_sample_bills_exist(self):
        assert len(SAMPLE_BILLS) >= 3

    def test_bill_numbers_unique(self):
        nums = [b[0] for b in SAMPLE_BILLS]
        assert len(nums) == len(set(nums))

    def test_bill_balance_due(self):
        for bill in SAMPLE_BILLS:
            total    = bill[9]
            payments = bill[10]
            balance  = bill[11]
            assert abs(balance - (total - payments)) < 0.01

    def test_bill_statuses(self):
        statuses = {b[12] for b in SAMPLE_BILLS}
        assert statuses <= {"Open", "Paid", "Overdue", "Void"}

    def test_bill_sheet_header(self, wb):
        ws = wb["Bills"]
        assert ws.cell(row=3, column=1).value == "Bill #"


# ---------------------------------------------------------------------------
# AP Payments
# ---------------------------------------------------------------------------

class TestAPPayments:
    def test_ap_pmt_headers_count(self):
        assert len(AP_PMT_HEADERS) == 10

    def test_sample_ap_pmts_exist(self):
        assert len(SAMPLE_AP_PMTS) >= 1

    def test_ap_pmt_sheet(self, wb):
        ws = wb["AP_Payments"]
        assert ws.cell(row=3, column=1).value == "Payment #"


# ---------------------------------------------------------------------------
# Journal Entries
# ---------------------------------------------------------------------------

class TestJournalEntries:
    def test_je_headers_count(self):
        assert len(JE_HEADERS) == 11

    def test_je_entries_exist(self):
        assert len(SAMPLE_JES) > 0

    def test_je_debits_equal_credits_per_entry(self):
        """For each journal entry number, total debits must equal total credits."""
        by_entry: dict = {}
        for je in SAMPLE_JES:
            entry_num = je[0]
            debit     = je[7]
            credit    = je[8]
            if entry_num not in by_entry:
                by_entry[entry_num] = [0.0, 0.0]
            by_entry[entry_num][0] += debit
            by_entry[entry_num][1] += credit

        for entry_num, (debits, credits) in by_entry.items():
            assert abs(debits - credits) < 0.01, (
                f"Entry {entry_num}: debits {debits:.2f} != credits {credits:.2f}"
            )

    def test_je_sheet_header(self, wb):
        ws = wb["Journal_Entries"]
        assert ws.cell(row=3, column=1).value == "Entry #"

    def test_je_sheet_has_data(self, wb):
        ws = wb["Journal_Entries"]
        assert ws.max_row >= 3 + len(SAMPLE_JES)


# ---------------------------------------------------------------------------
# Financial Report Sheets
# ---------------------------------------------------------------------------

class TestFinancialReports:
    def test_trial_balance_sheet_exists(self, wb):
        assert "Trial_Balance" in wb.sheetnames

    def test_trial_balance_has_accounts(self, wb):
        ws = wb["Trial_Balance"]
        # COA data rows start at row 5
        assert ws.max_row >= 5 + len(COA_DATA)

    def test_pl_sheet_exists(self, wb):
        assert "P&L" in wb.sheetnames

    def test_pl_has_revenue_section(self, wb):
        ws = wb["P&L"]
        # Row 4 should contain "REVENUE" section title
        found = any(
            ws.cell(row=r, column=1).value == "REVENUE"
            for r in range(3, 10)
        )
        assert found, "P&L sheet missing REVENUE section"

    def test_balance_sheet_exists(self, wb):
        assert "Balance_Sheet" in wb.sheetnames

    def test_balance_sheet_has_assets(self, wb):
        ws = wb["Balance_Sheet"]
        found = any(
            ws.cell(row=r, column=1).value == "ASSETS"
            for r in range(3, 10)
        )
        assert found, "Balance Sheet missing ASSETS section"

    def test_cash_flow_sheet_exists(self, wb):
        assert "Cash_Flow" in wb.sheetnames

    def test_cash_flow_has_operating_section(self, wb):
        ws = wb["Cash_Flow"]
        found = any(
            "Operating" in str(ws.cell(row=r, column=1).value or "")
            for r in range(3, 15)
        )
        assert found, "Cash Flow missing Operating Activities section"


# ---------------------------------------------------------------------------
# Aging Reports
# ---------------------------------------------------------------------------

class TestAgingReports:
    def test_ar_aging_sheet_exists(self, wb):
        assert "AR_Aging" in wb.sheetnames

    def test_ar_aging_has_header(self, wb):
        ws = wb["AR_Aging"]
        assert ws.cell(row=3, column=1).value == "Invoice #"
        # Verify standard 5-bucket aging columns
        headers = [ws.cell(row=3, column=c).value for c in range(1, 12)]
        assert "Current" in headers
        assert "1-30 Days" in headers
        assert "31-60 Days" in headers
        assert "61-90 Days" in headers
        assert "91+ Days" in headers

    def test_ap_aging_sheet_exists(self, wb):
        assert "AP_Aging" in wb.sheetnames

    def test_ap_aging_has_header(self, wb):
        ws = wb["AP_Aging"]
        assert ws.cell(row=3, column=1).value == "Bill #"
        # Verify standard 5-bucket aging columns
        headers = [ws.cell(row=3, column=c).value for c in range(1, 12)]
        assert "Current" in headers
        assert "1-30 Days" in headers
        assert "31-60 Days" in headers
        assert "61-90 Days" in headers
        assert "91+ Days" in headers


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------

class TestSettings:
    def test_settings_sheet_exists(self, wb):
        assert "Settings" in wb.sheetnames

    def test_settings_has_accounting_basis(self, wb):
        ws = wb["Settings"]
        # Check that "Accounting Basis" label is somewhere in the sheet
        found = any(
            "Accounting Basis" in str(ws.cell(row=r, column=1).value or "")
            for r in range(1, ws.max_row + 1)
        )
        assert found, "Settings sheet missing 'Accounting Basis' field"


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

class TestDashboard:
    def test_dashboard_exists(self, wb):
        assert "Dashboard" in wb.sheetnames

    def test_dashboard_has_title(self, wb):
        ws = wb["Dashboard"]
        title = ws.cell(row=1, column=1).value
        assert title is not None
        assert "ProBooksAi" in str(title)


# ---------------------------------------------------------------------------
# File output
# ---------------------------------------------------------------------------

class TestFileOutput:
    def test_file_created(self, workbook_path):
        assert os.path.exists(workbook_path)

    def test_file_not_empty(self, workbook_path):
        assert os.path.getsize(workbook_path) > 10_000  # at least 10 KB

    def test_file_is_valid_xlsx(self, workbook_path):
        """File must be loadable as a valid Excel workbook."""
        wb = load_workbook(workbook_path)
        assert len(wb.sheetnames) > 0
